import proTeams

awayTeamRowInfo=proTeams.findTeamRowInTuple("BOS") #team name abbrev in, returns "BOS","Boston Bruins","BOS.png"
print(awayTeamRowInfo)

#############
import requests
import json
import os
import sys
import proTeams
# import teamGameCount
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime, timedelta, date
from PIL import Image

#This class parses thru the API json and creates a text file of the info for the schedule
class GetNHLSchedule:
    def __init__(self,weekDate):
        weekDateCall='https://api-web.nhle.com/v1/schedule/'+weekDate #Looks like: 'https://api-web.nhle.com/v1/schedule/2024-03-25'
        self.nhlApi=requests.get(weekDateCall)

    def delete_file_if_exists(self, fname): #check for file, delete if exists to avoid duplicate schedules
        if os.path.exists(fname): # Check if the file exists
            os.remove(fname) # If it exists, delete the file
            print(f"File '{fname}' deleted.")
        else:
            print(f"File '{fname}' does not exist.")

    def getNhlGameInfo(self, fname):
        r=self.nhlApi
        if r.status_code != 200:
            print ("Problem connecting with NHL API")
            exit
        self.delete_file_if_exists(fname) #delete the file if it already exists so we don't append to it
        f = open(fname, "a")
        theJSON = json.loads(r.content)
        for i in theJSON["gameWeek"]:
            dateGameOfWeek = i["date"]
            dayAbbrev = i["dayAbbrev"]
            dayNumberOfGames = i["numberOfGames"]
            for aRow in i["games"]:
                awayTeamName = aRow["awayTeam"]["placeName"]["default"]
                awayTeamAbbrev = aRow["awayTeam"]["abbrev"]
                homeTeamName = aRow["homeTeam"]["placeName"]["default"]
                homeTeamAbbrev = aRow["homeTeam"]["abbrev"]
                gameInfo=dateGameOfWeek+","+dayAbbrev+","+str(dayNumberOfGames)+","+awayTeamAbbrev+","+homeTeamAbbrev+'\n'
                f.write(gameInfo)
        f.close()

#A class to create the excel file with the scheduled data
class WriteNHLSchedule:
    def __init__(self, xName):
        self.filename = xName
        self.workbook = Workbook()
        self.ws = self.workbook.active

    def set_row_height(self, row, height):
        self.ws.row_dimensions[row].height = height

    def set_column_width(self, column, width):
        self.ws.column_dimensions[column].width = width

    def set_cell_font(self, row, column, font_name='Arial', font_size=11, bold=False, color=None):
        cell = self.ws.cell(row=row, column=column)
        cell.font = Font(name=font_name, size=font_size, bold=bold, color=color)

    def set_cell_alignment(self, row, column, horizontal='center', vertical='center'):
        cell = self.ws.cell(row=row, column=column)
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

    def set_cell_border(self, row, column, border_style='thin'):
        cell = self.ws.cell(row=row, column=column)
        border = Border(left=Side(style=border_style), right=Side(style=border_style), 
                        top=Side(style=border_style), bottom=Side(style=border_style))
        cell.border = border

    def set_cell_fill_color(self, row, column, color='FFFFFF'):
        cell = self.ws.cell(row=row, column=column)
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.fill = fill

    def set_column_width(self, start_column, end_column, width):
        for col in range(start_column, end_column):
            self.ws.column_dimensions[self.ws.cell(row=1, column=col).column_letter].width = width

    def write_row_data(self, row, data):
        for i, value in enumerate(data, start=1):
            self.ws.cell(row=row, column=i, value=value)

    def write_column_data(self, row, column, value):
        self.ws.cell(row=row, column=column, value=value)

    def insert_team_logo(self, row, column, imageName):
        row=str(row)
        cellName=column+row
        from PIL import Image as PILImage #Pillow (PIL) was needed for dealing with images
        pil_image = PILImage.open(imageName)
        excel_image = ExcelImage(pil_image)
        self.ws.add_image(excel_image, cellName)

    def save_excel(self):
        self.workbook.save(self.filename)

    def close_excel(self):
        self.workbook.close(self.filename)

#######
#Beginning of main code

#Calling a class to parse thru the API json and creates a text file of the info for the schedule
dateForTheWeek='2024-04-08'
a=GetNHLSchedule(dateForTheWeek)
filename = "c:\\Temp\\demoHockeySchedule.txt"
a.getNhlGameInfo(filename)

#Calling a class to create the excel file with the scheduled data
xName="c:\\Temp\\hockeydemo.xlsx"
rowOneHeader = ["Abbrev","Team Logo","MON","TUE","WED","THU","FRI","SAT","SUN","Game Count"]
rowOneHeaderCount=len(rowOneHeader)
excelNhlSchedule=WriteNHLSchedule(xName)

#Setup left columns for NHL teams
imagePath="C:\\Temp\\HockeyTeamLogos\\"
# imageName=imagePath+"images\\BOS.png"
imageName="./images/BOS.png"
excelNhlSchedule.insert_team_logo(1, "B", imageName)
excelNhlSchedule.save_excel()
sys.exit()

row=4
for i, value in enumerate(proTeams.proTeamTuple, start=00):
    excelNhlSchedule.set_row_height(row, 35)
    excelNhlSchedule.set_cell_alignment(row, 1, horizontal='center', vertical='center')
    excelNhlSchedule.set_cell_font(row, 1, font_name="Georgia", bold=True, color='367ee7')
    excelNhlSchedule.write_column_data(row, 1, value[0]) #Can use numbers too for column name
    imageName=imagePath+value[2]
    excelNhlSchedule.insert_team_logo(row, "B", imageName)
    row+=1

#Cell color fill for the empty cells not used
cellFill=[1,2,10]
for x in range(2, 4):
    for i, value in enumerate(cellFill, start=0):
        excelNhlSchedule.set_cell_fill_color(x, cellFill[i], color='c0c0c0')

#Write the schedule now. Find the matching team and column to write to
#Look up the teams and fill it in on the spreadsheet - basically twice for each team
rowOffset=4 #to get positioned, first team in spreadsheet starts at row 4
col=2
colLetter='B'
firstRetrieval="True"
f = open(filename,'r')
i=0
while True:
        x=f.readline()
        if not x:
            print("EOF reached")
            f.close()
            break
        #Parse line: date, day, nbr of games, away, home
        gameInfo= x.split(",") #2024-03-25,MON,2,VGK,STL
        #These if statements are to control the logic of what column to write to
        if firstRetrieval=="True":
            firstRetrieval="False"
            currentGameDate=gameInfo[0]
            nextGameDate=currentGameDate
            i+=1
            col+=1
            colLetter=chr (ord (colLetter) + 1) #columns are letters, increment the column to write to the correct one starting with C
            excelNhlSchedule.set_cell_font(3, col, bold=True, color='17020e')
            excelNhlSchedule.set_cell_alignment(3, col, horizontal='center', vertical='center')
            excelNhlSchedule.set_cell_fill_color(3, col, color='0fcbfd') 
            excelNhlSchedule.write_column_data(3, col, gameInfo[2])
        else:
            if i<=int(gameInfo[2]):
                i+=1
                nextGameDate=gameInfo[0]
        
        if currentGameDate!=nextGameDate:
            col+=1
            colLetter=chr (ord (colLetter) + 1) #columns are letters, increment the column to write to the correct one starting with C

        #Get info setup 
        awayTeam=gameInfo[3]
        homeTeam=gameInfo[4]
        homeTeam=homeTeam[0:3] #returning the first 3 characters to avoid the \n
        awayTeamRowInfo=proTeams.findTeamRowInTuple(awayTeam) #team name abbrev in, returns "BOS","Boston Bruins","BOS.png"
        if awayTeamRowInfo == None: #exit when no team was found; probably a typo
            print("No Away team was found, exiting program as there is a problem; possibly a typo on team name.")
            sys.exit()
        homeTeamRowInfo=proTeams.findTeamRowInTuple(homeTeam) #team name abbrev in, returns "BOS","Boston Bruins","BOS.png"
        if homeTeamRowInfo == None: #exit when no team was found; probably a typo
            print("No Home team was found, exiting program as there is a problem; possibly a typo on team name.")
            sys.exit()
        #Process the matchups - first the Away Team appears in the list
        #Find the Away Team POSITION in the spreadsheet, then write/insert the HOME Team logo
        #first team is always Away Team gameInfo[3]
        indexOfTeam=proTeams.findIndexOfTeamInTuple(awayTeam) #using this to figure out where in the spreadsheet the teams are located
        rowPosition=indexOfTeam+rowOffset
        excelNhlSchedule.set_cell_alignment(rowPosition, col, horizontal='center', vertical='center')
        excelNhlSchedule.set_cell_fill_color(rowPosition, col, color='4edc9c')
        #Now insert the Home Team image
        imageName=imagePath+homeTeamRowInfo[2] #name of team image
        excelNhlSchedule.insert_team_logo(rowPosition, colLetter, imageName)
        teamGameCount.teamGameCountIncrement(awayTeam) 
        #Now doing the other combo of the matchup
        indexOfTeam=proTeams.findIndexOfTeamInTuple(homeTeam) #using this to figure out where in the spreadsheet the teams are located
        rowPosition=indexOfTeam+rowOffset
        excelNhlSchedule.set_cell_alignment(rowPosition, col, horizontal='center', vertical='center')
        excelNhlSchedule.set_cell_fill_color(rowPosition, col, color='ff1919')
        #Now insert the Away Team image
        imageName=imagePath+awayTeamRowInfo[2] #name of team image
        excelNhlSchedule.insert_team_logo(rowPosition, colLetter, imageName)
        teamGameCount.teamGameCountIncrement(homeTeam)
        if i==int(gameInfo[2]):
            i=0
            firstRetrieval="True"
        elif i>int(gameInfo[2]):
            sys.exit("Problem with date count/comparison, got larger, exiting out.")

for i, value in enumerate(proTeams.proTeamTuple, start=00):
    indexOfTeam=proTeams.findIndexOfTeamInTuple(value[0]) #using this to figure out where in the spreadsheet the teams are located
    rowPosition=indexOfTeam+rowOffset
    gameCount=teamGameCount.teamGameCountRetrieval(value[0])
    excelNhlSchedule.set_cell_font(rowPosition, 10, bold=True, color='17020e')
    excelNhlSchedule.set_cell_alignment(rowPosition, 10, horizontal='center', vertical='center')
    excelNhlSchedule.set_cell_fill_color(rowPosition, 10, color='0fcbfd')
    excelNhlSchedule.write_column_data(rowPosition, 10, gameCount)

excelNhlSchedule.save_excel()
